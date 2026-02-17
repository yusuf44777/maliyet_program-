"""
Maliyet Sistemi - Excel Motor Modülü
maliyet_sablonu.xlsx formatında okuma/yazma işlemleri.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import os
import re

from storage_utils import is_http_url, cache_remote_file

TOKEN_PATTERN = re.compile(r"[a-z0-9çğıöşü]+", re.I)
SILVER_TOKENS = {"silver", "gumus", "gümüş", "gümus"}
GOLD_COPPER_TOKENS = {
    "gold", "altin", "altın",
    "copper", "bakir", "bakır",
    "bronze", "pirinc", "pirinç",
    "rosegold",
}


def resolve_template_path() -> Path:
    """
    Kullanılacak şablon dosyasını belirler.
    Öncelik:
    1) TEMPLATE_PATH env
    2) kategori_calismasi/en_son_maliyet_sablonu.xlsx
    3) maliyet_programı/en_son_maliyet_sablonu.xlsx
    4) kategori_calismasi/son_maliyet_sablonu.xlsx
    5) maliyet_programı/son_maliyet_sablonu.xlsx
    6) maliyet_programı/maliyet_sablonu.xlsx
    """
    env_path = os.getenv("TEMPLATE_PATH", "").strip()
    env_url = os.getenv("TEMPLATE_URL", "").strip()
    candidates = []

    template_url = env_url or (env_path if is_http_url(env_path) else "")
    if template_url:
        cached = cache_remote_file(
            template_url,
            cache_name="template.xlsx",
            ttl_seconds=int(os.getenv("REMOTE_FILE_CACHE_TTL", "900")),
        )
        if cached and cached.exists():
            return cached

    if env_path and not is_http_url(env_path):
        candidates.append(Path(env_path).expanduser())

    base_dir = Path(__file__).resolve().parent.parent
    root_dir = Path(__file__).resolve().parents[2]
    candidates.extend([
        root_dir / "en_son_maliyet_sablonu.xlsx",
        base_dir / "en_son_maliyet_sablonu.xlsx",
        root_dir / "son_maliyet_sablonu.xlsx",
        base_dir / "son_maliyet_sablonu.xlsx",
        base_dir / "maliyet_sablonu.xlsx",
    ])

    for p in candidates:
        if p.exists():
            return p
    return candidates[-1]


def normalize_text(value: str | None) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("ı", "i").replace("İ", "i")
    text = text.replace("’", "'").replace("`", "'")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*([,/()*-])\s*", r"\1", text)
    return text.strip()


def tokenize_text(value: str | None) -> set[str]:
    if not value:
        return set()
    return {t for t in TOKEN_PATTERN.findall(str(value).lower()) if len(t) > 1}


def detect_kaplama_tier(*values: str | None) -> str:
    tokens: set[str] = set()
    for value in values:
        tokens.update(tokenize_text(value))
    if tokens & GOLD_COPPER_TOKENS:
        return "gold_copper"
    if tokens & SILVER_TOKENS:
        return "silver"
    return "other"


def split_cost_base_and_tier(cost_name: str | None) -> tuple[str, str]:
    raw = str(cost_name or "").strip()
    if not raw:
        return "", "other"
    m = re.match(r"^(.*?)\s*\(([^)]*)\)\s*$", raw)
    if not m:
        return raw, "other"
    suffix = m.group(2)
    tier = detect_kaplama_tier(suffix)
    if tier == "other":
        return raw, "other"
    return m.group(1).strip(), tier


def get_template_structure():
    """
    Şablonun kolon yapısını analiz eder.
    Returns: dict with 'info_cols', 'cost_cols', 'material_cols'
    """
    template_path = resolve_template_path()
    wb = openpyxl.load_workbook(template_path, read_only=True)
    ws = wb["Maliyet Şablonu"]

    structure = {
        "info_cols": {},       # A-G: Ürün bilgileri
        "cost_cols": {},       # H-DW: Maliyet (ambalaj) kolonları
        "material_cols": {},   # DX-EX: Hammadde kolonları
    }

    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if not header:
            continue

        header_str = str(header)
        col_letter = get_column_letter(col)

        if header_str.startswith("Maliyet:"):
            name = header_str.replace("Maliyet:", "").strip()
            structure["cost_cols"][col] = {
                "letter": col_letter,
                "header": header_str,
                "name": name,
            }
        elif header_str.startswith("Hammadde:"):
            raw = header_str.replace("Hammadde:", "").strip()
            unit_match = re.search(r"\(([^)]+)\)", raw)
            unit = unit_match.group(1) if unit_match else "pcs"
            name = re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()
            structure["material_cols"][col] = {
                "letter": col_letter,
                "header": header_str,
                "name": name,
                "unit": unit,
            }
        else:
            structure["info_cols"][col] = {
                "letter": col_letter,
                "header": header_str,
            }

    wb.close()
    return structure


def export_to_template(products: list[dict], output_path: str = None) -> str:
    """
    Ürünleri maliyet_sablonu formatında Excel'e yazar.
    
    products: Her biri şu anahtarları taşıyabilir:
        - child_sku, child_name, en, boy, alan_m2
        - materials: {material_name: quantity}
        - costs: {cost_name: 'x'}
    
    Returns: Oluşturulan dosya yolu
    """
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = Path(__file__).parent.parent / "exports"
        output_dir.mkdir(exist_ok=True)
        output_path = str(output_dir / f"maliyet_export_{timestamp}.xlsx")

    template_path = resolve_template_path()

    # Şablonu kopyala (sadece header satırını al)
    template_wb = openpyxl.load_workbook(template_path, read_only=True)
    template_ws = template_wb["Maliyet Şablonu"]

    # Yeni workbook oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maliyet Şablonu"

    # Header satırını kopyala
    headers = {}
    for col in range(1, template_ws.max_column + 1):
        val = template_ws.cell(row=1, column=col).value
        if val:
            ws.cell(row=1, column=col, value=val)
            headers[str(val)] = col

    template_wb.close()

    # Kolon indekslerini belirle
    structure = get_template_structure()

    # Ürün bilgi alanlarının kolon haritası
    info_map = {
        "Ürün Kodu": None,
        "Ürün Adı": None,
        "En": None,
        "Boy": None,
        "Yükseklik": None,
        "Ağırlık": None,
        "Desi": None,
    }
    for col_idx, info in structure["info_cols"].items():
        if info["header"] in info_map:
            info_map[info["header"]] = col_idx

    # Hammadde kolon haritası
    material_col_map: dict[str, int] = {}
    material_col_map_norm: dict[str, int] = {}
    for col_idx, mat_info in structure["material_cols"].items():
        mat_name = mat_info["name"]
        material_col_map[mat_name] = col_idx
        material_col_map_norm[normalize_text(mat_name)] = col_idx

    # Maliyet kolon haritası (exact + normalize + base/tier)
    cost_col_map: dict[str, int] = {}
    cost_col_map_norm: dict[str, int] = {}
    cost_base_candidates: dict[str, list[dict]] = {}
    for col_idx, cost_info in structure["cost_cols"].items():
        cost_name = cost_info["name"]
        cost_col_map[cost_name] = col_idx
        cost_col_map_norm[normalize_text(cost_name)] = col_idx

        base_name, tier = split_cost_base_and_tier(cost_name)
        base_key = normalize_text(base_name)
        cost_base_candidates.setdefault(base_key, []).append({
            "col": col_idx,
            "tier": tier,
            "name": cost_name,
        })

    def resolve_material_col(material_name: str) -> int | None:
        if material_name in material_col_map:
            return material_col_map[material_name]
        return material_col_map_norm.get(normalize_text(material_name))

    def resolve_cost_col(cost_name: str, product: dict) -> int | None:
        # 1) birebir eşleşme
        if cost_name in cost_col_map:
            return cost_col_map[cost_name]

        # 2) normalize eşleşme
        norm_key = normalize_text(cost_name)
        if norm_key in cost_col_map_norm:
            return cost_col_map_norm[norm_key]

        # 3) base-name fallback (özellikle yeni şablondaki silver/gold,copper ayrımı için)
        requested_base, requested_tier = split_cost_base_and_tier(cost_name)
        base_key = normalize_text(requested_base or cost_name)
        candidates = cost_base_candidates.get(base_key, [])
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]["col"]

        # Çoklu eşleşmede ürün rengine göre tier seç
        product_tier = requested_tier if requested_tier != "other" else detect_kaplama_tier(
            product.get("variation_color"),
            product.get("child_name"),
        )
        same_tier = [c for c in candidates if c["tier"] == product_tier]
        if same_tier:
            return same_tier[0]["col"]

        # Son fallback: other tier veya ilk kolon
        other = [c for c in candidates if c["tier"] == "other"]
        if other:
            return other[0]["col"]
        return candidates[0]["col"]

    # Ürünleri yaz
    for row_idx, product in enumerate(products, start=2):
        # Temel bilgiler
        if info_map["Ürün Kodu"]:
            ws.cell(row=row_idx, column=info_map["Ürün Kodu"],
                    value=product.get("child_sku"))
        if info_map["Ürün Adı"]:
            ws.cell(row=row_idx, column=info_map["Ürün Adı"],
                    value=product.get("child_name"))
        if info_map["En"]:
            ws.cell(row=row_idx, column=info_map["En"],
                    value=product.get("en"))
        if info_map["Boy"]:
            ws.cell(row=row_idx, column=info_map["Boy"],
                    value=product.get("boy"))
        if info_map["Yükseklik"]:
            ws.cell(row=row_idx, column=info_map["Yükseklik"],
                    value=product.get("yukseklik"))
        if info_map["Ağırlık"]:
            ws.cell(row=row_idx, column=info_map["Ağırlık"],
                    value=product.get("agirlik"))
        if info_map["Desi"]:
            ws.cell(row=row_idx, column=info_map["Desi"],
                    value=product.get("desi"))

        # Hammaddeler
        materials = product.get("materials", {})
        for mat_name, quantity in materials.items():
            if not quantity:
                continue
            mat_col = resolve_material_col(mat_name)
            if mat_col:
                ws.cell(row=row_idx, column=mat_col, value=quantity)

        # Maliyet (ambalaj) atamaları
        costs = product.get("costs", {})
        for cost_name, marker in costs.items():
            cost_col = resolve_cost_col(cost_name, product)
            if cost_col:
                ws.cell(row=row_idx, column=cost_col, value=marker)

    wb.save(output_path)
    return output_path


def read_template_products() -> list[dict]:
    """
    Ana şablondaki tüm ürünlerin temel bilgilerini okur (Ürün Kodu, Adı, En, Boy).
    """
    template_path = resolve_template_path()
    wb = openpyxl.load_workbook(template_path, read_only=True)
    ws = wb["Maliyet Şablonu"]

    products = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        sku = row[0].value  # A kolonu
        if not sku:
            continue
        products.append({
            "child_sku": str(sku),
            "child_name": row[1].value,  # B
            "en": row[2].value,           # C
            "boy": row[3].value,          # D
            "yukseklik": row[4].value,    # E
            "agirlik": row[5].value,      # F
            "desi": row[6].value,         # G
        })

    wb.close()
    return products
