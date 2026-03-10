"""
Maliyet Sistemi - Veritabanı Modülü
mapped_products CSV'lerini DB'ye yükler ve alan hesaplamalarını yapar.
Yalnızca PostgreSQL/Supabase destekler.
"""

import csv
import math
import re
import os
import threading
import logging
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor, execute_batch as PgExecuteBatch
    from psycopg2 import IntegrityError as PgIntegrityError
    from psycopg2.pool import ThreadedConnectionPool
except Exception:
    psycopg2 = None
    RealDictCursor = None
    PgExecuteBatch = None
    PgIntegrityError = None
    ThreadedConnectionPool = None

# Vercel ortamında psycopg2 yoksa psycopg (v3) compat katmanını dene
if psycopg2 is None:
    try:
        from psycopg import connect as _psycopg3_connect
        from psycopg.rows import dict_row as _psycopg3_dict_row
        from psycopg import errors as _psycopg3_errors

        class _Psycopg3Compat:
            """psycopg2 API'sine benzer ince wrapper — sadece kullandığımız kısımlar."""
            @staticmethod
            def connect(dsn):
                conn = _psycopg3_connect(dsn, row_factory=_psycopg3_dict_row, autocommit=False)
                return conn

        psycopg2 = _Psycopg3Compat()
        RealDictCursor = None  # psycopg3'te dict_row zaten row_factory olarak set ediliyor
        PgExecuteBatch = None
        PgIntegrityError = _psycopg3_errors.IntegrityError
    except Exception:
        pass

from storage_utils import is_http_url, cache_remote_file

logger = logging.getLogger("maliyet.db")

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
if not DATABASE_URL.startswith(("postgres://", "postgresql://")):
    raise RuntimeError(
        "DATABASE_URL zorunlu ve PostgreSQL formatında olmalı. "
        "Örnek: postgresql://USER:PASSWORD@HOST:PORT/DB?sslmode=require"
    )


def _parse_database_url_metadata(url: str) -> dict[str, Any]:
    try:
        parsed = urlparse(url)
        host = (parsed.hostname or "").strip().lower()
        port = parsed.port
    except Exception:
        host = ""
        port = None

    is_supabase_pooler = host.endswith(".pooler.supabase.com")
    is_supabase_direct = host.startswith("db.") and host.endswith(".supabase.co")
    is_supabase = is_supabase_pooler or host.endswith(".supabase.co")

    if port is None:
        if is_supabase_pooler:
            port = 6543
        elif is_supabase_direct:
            port = 5432

    if is_supabase_pooler:
        url_kind = "supabase_pooler"
    elif is_supabase_direct:
        url_kind = "supabase_direct"
    else:
        url_kind = "postgres"

    return {
        "host": host,
        "port": port,
        "is_supabase": is_supabase,
        "is_supabase_pooler": is_supabase_pooler,
        "is_supabase_direct": is_supabase_direct,
        "url_kind": url_kind,
    }


_database_url_metadata = _parse_database_url_metadata(DATABASE_URL)
DATABASE_HOST = _database_url_metadata["host"]
DATABASE_PORT = _database_url_metadata["port"]
DATABASE_IS_SUPABASE = _database_url_metadata["is_supabase"]
DATABASE_IS_SUPABASE_POOLER = _database_url_metadata["is_supabase_pooler"]
DATABASE_IS_SUPABASE_DIRECT = _database_url_metadata["is_supabase_direct"]
DATABASE_URL_KIND = _database_url_metadata["url_kind"]
DB_BACKEND = "postgres"
IS_POSTGRES = True
IntegrityError = PgIntegrityError if PgIntegrityError is not None else Exception


def _env_flag(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}


PG_POOL_ENABLED = _env_flag("PG_POOL_ENABLED", default=True)
PG_POOL_MIN_CONN = max(1, int(os.getenv("PG_POOL_MIN_CONN", "1")))
PG_POOL_MAX_CONN = max(PG_POOL_MIN_CONN, int(os.getenv("PG_POOL_MAX_CONN", "3")))
PG_CONNECT_TIMEOUT = max(2, int(os.getenv("PG_CONNECT_TIMEOUT", "10")))
PG_EXECUTEMANY_PAGE_SIZE = max(50, int(os.getenv("PG_EXECUTEMANY_PAGE_SIZE", "500")))
_pg_pool = None
_pg_pool_lock = threading.Lock()


def _get_or_create_pg_pool():
    """
    Supabase pooler önünde bile local process içinde bağlantı reuse etmek
    serverless connection spike'larını azaltır.
    """
    global _pg_pool
    if not (PG_POOL_ENABLED and ThreadedConnectionPool is not None):
        return None
    if _pg_pool is not None:
        return _pg_pool
    with _pg_pool_lock:
        if _pg_pool is None:
            _pg_pool = ThreadedConnectionPool(
                minconn=PG_POOL_MIN_CONN,
                maxconn=PG_POOL_MAX_CONN,
                dsn=DATABASE_URL,
                connect_timeout=PG_CONNECT_TIMEOUT,
            )
    return _pg_pool


def close_pg_pool():
    """İhtiyaç halinde tüm pooled bağlantıları kapatır."""
    global _pg_pool
    with _pg_pool_lock:
        pool = _pg_pool
        _pg_pool = None
    if pool is not None:
        pool.closeall()


def _reset_pg_pool():
    """Bozuk/stale pool'u tamamen kapatıp sıfırlar."""
    close_pg_pool()


def _acquire_healthy_pooled_conn(pool):
    """
    Havuzdan bağlantı alırken stale/broken bağlantıyı eleyip yeniden dener.
    """
    last_error = None
    for _ in range(2):
        raw_conn = None
        try:
            raw_conn = pool.getconn()
            cur = raw_conn.cursor()
            try:
                cur.execute("SELECT 1")
                _ = cur.fetchone()
            finally:
                try:
                    cur.close()
                except Exception:
                    pass
            return raw_conn
        except Exception as exc:
            last_error = exc
            try:
                pool.putconn(raw_conn, close=True)
            except TypeError:
                try:
                    raw_conn.close()
                except Exception:
                    pass
            except Exception:
                pass
    if last_error:
        raise last_error
    raise RuntimeError("PostgreSQL connection pool: sağlıklı bağlantı alınamadı")


def _create_raw_connection():
    try:
        return psycopg2.connect(DATABASE_URL, connect_timeout=PG_CONNECT_TIMEOUT)
    except TypeError:
        # psycopg3 compat wrapper sadece dsn parametresi alıyor.
        return psycopg2.connect(DATABASE_URL)


def _set_connection_autocommit(raw_conn):
    try:
        raw_conn.autocommit = False
    except Exception:
        pass
    return raw_conn


def get_database_diagnostics() -> dict[str, Any]:
    return {
        "database_url_kind": DATABASE_URL_KIND,
        "database_port": DATABASE_PORT,
        "database_uses_supabase": DATABASE_IS_SUPABASE,
        "database_uses_supabase_pooler": DATABASE_IS_SUPABASE_POOLER,
        "database_uses_supabase_direct": DATABASE_IS_SUPABASE_DIRECT,
        "pg_pool_enabled": bool(PG_POOL_ENABLED and ThreadedConnectionPool is not None),
    }


PROJECT_ROOT = Path(__file__).resolve().parent.parent  # maliyet_programı root
KATEGORI_ROOT = PROJECT_ROOT.parent  # kategori_calismasi root

KATEGORI_DIRS = {
    "metal": KATEGORI_ROOT / "metal_kategori",
    "ahsap": KATEGORI_ROOT / "ahsap_kategori",
    "cam": KATEGORI_ROOT / "cam_kategori",
    "harita": KATEGORI_ROOT / "harita_kategori",
    "mobilya": KATEGORI_ROOT / "mobilya_kategori",
}

KATEGORI_ALIASES = {
    "ahsap": "ahsap",
    "ahşap": "ahsap",
    "mobilya": "mobilya",
    "metal": "metal",
    "cam": "cam",
    "harita": "harita",
}


PRODUCT_EXTRA_COLUMNS = [
    ("kargo_kodu", "TEXT"),
    ("kargo_en", "REAL"),
    ("kargo_boy", "REAL"),
    ("kargo_yukseklik", "REAL"),
    ("kargo_agirlik", "REAL"),
    ("kargo_desi", "REAL"),
    ("ham_maliyet_status", "TEXT"),
    ("is_active", "INTEGER NOT NULL DEFAULT 1"),
]

KARGO_CODE_PATTERN = re.compile(r"([A-Z])\s*-\s*(\d+[A-Z]?)", re.I)
LEGACY_GOLD_SILVER_SUFFIX_PATTERN = re.compile(r"\(\s*gold\s*,\s*silver\s*\)\s*$", re.I)
KAPLAMA_COLOR_SUFFIX_PATTERN = re.compile(
    r"\(\s*(?:silver|gumus|gümüş|gümus|gold|altin|altın|copper|bakir|bakır|bronze|pirinc|pirinç|rosegold)"
    r"(?:\s*,\s*(?:silver|gumus|gümüş|gümus|gold|altin|altın|copper|bakir|bakır|bronze|pirinc|pirinç|rosegold))*\s*\)\s*$",
    re.I,
)

INSERT_OR_IGNORE_PATTERN = re.compile(r"^\s*INSERT\s+OR\s+IGNORE\s+INTO\s+", re.I)


def is_postgres_backend() -> bool:
    return IS_POSTGRES


def get_supported_categories() -> list[str]:
    return list(KATEGORI_DIRS.keys())


def normalize_product_categories(categories: list[str] | None = None) -> list[str]:
    """
    Girilen kategori listesini normalize eder ve desteklenenleri döndürür.
    Boş gelirse tüm kategoriler döner.
    """
    if not categories:
        return get_supported_categories()

    out: list[str] = []
    seen: set[str] = set()
    invalid: list[str] = []

    for raw in categories:
        key = str(raw or "").strip().lower()
        normalized = KATEGORI_ALIASES.get(key, key)
        if normalized not in KATEGORI_DIRS:
            invalid.append(str(raw))
            continue
        if normalized in seen:
            continue
        seen.add(normalized)
        out.append(normalized)

    if invalid:
        supported = ", ".join(get_supported_categories())
        raise ValueError(f"Desteklenmeyen kategori(ler): {', '.join(invalid)}. Desteklenenler: {supported}")
    return out


def adapt_sql_for_backend(sql: str) -> str:
    """
    SQLite-odaklı SQL'i PostgreSQL'e en az sürtünmeyle uyarlar.
    """
    query = str(sql)

    if INSERT_OR_IGNORE_PATTERN.match(query):
        query = INSERT_OR_IGNORE_PATTERN.sub("INSERT INTO ", query, count=1)
        trimmed = query.rstrip()
        has_semicolon = trimmed.endswith(";")
        if has_semicolon:
            trimmed = trimmed[:-1]
        query = f"{trimmed} ON CONFLICT DO NOTHING"
        if has_semicolon:
            query += ";"

    return query.replace("?", "%s")


def adapt_params(params: Any):
    if params is None:
        return ()
    if isinstance(params, list):
        return tuple(params)
    return params


class PGCompatCursor:
    def __init__(self, inner):
        self._inner = inner

    def execute(self, sql: str, params=None):
        query = adapt_sql_for_backend(sql)
        if params is None:
            self._inner.execute(query)
        else:
            self._inner.execute(query, adapt_params(params))
        return self

    def executemany(self, sql: str, seq_of_params):
        query = adapt_sql_for_backend(sql)
        adapted_params = [adapt_params(p) for p in (seq_of_params or [])]
        if not adapted_params:
            return self
        if IS_POSTGRES and PgExecuteBatch is not None:
            PgExecuteBatch(
                self._inner,
                query,
                adapted_params,
                page_size=PG_EXECUTEMANY_PAGE_SIZE,
            )
        else:
            self._inner.executemany(query, adapted_params)
        return self

    def fetchone(self):
        return self._inner.fetchone()

    def fetchall(self):
        return self._inner.fetchall()

    def __iter__(self):
        return iter(self._inner)

    def __getattr__(self, name):
        return getattr(self._inner, name)


class PGCompatConnection:
    def __init__(self, inner, pool=None):
        self._inner = inner
        self._pool = pool
        self._is_closed = False

    def cursor(self):
        if RealDictCursor is not None:
            return PGCompatCursor(self._inner.cursor(cursor_factory=RealDictCursor))
        return PGCompatCursor(self._inner.cursor())

    def execute(self, sql: str, params=None):
        cur = self.cursor()
        cur.execute(sql, params)
        return cur

    def executemany(self, sql: str, seq_of_params):
        cur = self.cursor()
        cur.executemany(sql, seq_of_params)
        return cur

    def commit(self):
        return self._inner.commit()

    def rollback(self):
        return self._inner.rollback()

    def close(self):
        if self._is_closed:
            return None
        if self._pool is not None:
            # Broken transaction state'in pool'a geri dönmesini engelle.
            broken = False
            try:
                self._inner.rollback()
            except Exception:
                broken = True
            try:
                self._pool.putconn(self._inner, close=broken)
            except TypeError:
                if broken:
                    try:
                        self._inner.close()
                    except Exception:
                        pass
                else:
                    self._pool.putconn(self._inner)
            except Exception:
                if broken:
                    try:
                        self._inner.close()
                    except Exception:
                        pass
        else:
            self._inner.close()
        self._is_closed = True
        return None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        if exc_type is not None:
            try:
                self.rollback()
            except Exception:
                pass
        self.close()
        return False

    def __del__(self):
        try:
            self.close()
        except Exception:
            pass

    def __getattr__(self, name):
        return getattr(self._inner, name)


def resolve_template_path() -> Path:
    """
    Kullanılacak şablon dosyasını belirler.
    Öncelik:
    1) TEMPLATE_PATH env
    2) kategori_calismasi/en_son_maliyet_sablonu.xlsx
    3) maliyet_programı/en_son_maliyet_sablonu.xlsx
    4) kategori_calismasi/son_maliyet_sablonu.xlsx
    5) maliyet_programı/son_maliyet_sablonu.xlsx
    6) maliyet_programı/maliyet_sablonu.xlsx
    """
    env_path = os.getenv("TEMPLATE_PATH", "").strip()
    env_url = os.getenv("TEMPLATE_URL", "").strip()
    candidates = []

    template_url = env_url or (env_path if is_http_url(env_path) else "")
    if template_url:
        cached = cache_remote_file(
            template_url,
            cache_name="template.xlsx",
            ttl_seconds=int(os.getenv("REMOTE_FILE_CACHE_TTL", "900")),
        )
        if cached and cached.exists():
            return cached

    if env_path and not is_http_url(env_path):
        candidates.append(Path(env_path).expanduser())

    base_dir = Path(__file__).resolve().parent.parent
    root_dir = Path(__file__).resolve().parents[2]
    candidates.extend([
        root_dir / "en_son_maliyet_sablonu.xlsx",
        base_dir / "en_son_maliyet_sablonu.xlsx",
        root_dir / "son_maliyet_sablonu.xlsx",
        base_dir / "son_maliyet_sablonu.xlsx",
        base_dir / "maliyet_sablonu.xlsx",
    ])

    for p in candidates:
        if p.exists():
            return p
    return candidates[-1]


def canonicalize_kaplama_cost_name(name: str | None) -> str:
    """
    Kaplama maliyet adlarını tek forma indirger.
    Legacy: '(gold,silver)' -> '(gold,copper)'
    """
    raw = str(name or "").strip()
    if not raw:
        return ""
    return LEGACY_GOLD_SILVER_SUFFIX_PATTERN.sub("(gold,copper)", raw)


def split_kaplama_tier_suffix(name: str | None) -> tuple[str, str | None]:
    """KAPLAMA_COLOR_SUFFIX_PATTERN ile base ad + tier suffix ayrıştırır."""
    raw = str(name or "").strip()
    if not raw:
        return "", None
    m = KAPLAMA_COLOR_SUFFIX_PATTERN.search(raw)
    if not m:
        return raw, None
    base = raw[:m.start()].strip()
    suffix = m.group(0).strip()[1:-1].strip().lower()
    return base, suffix


def get_db():
    """PostgreSQL bağlantısı döndürür."""
    if psycopg2 is None:
        raise RuntimeError("PostgreSQL driver yüklenemedi. 'psycopg2-binary' kurulmalı.")

    last_error = None

    if PG_POOL_ENABLED and ThreadedConnectionPool is not None:
        for attempt in range(2):
            try:
                pool = _get_or_create_pg_pool()
                if pool is None:
                    break
                raw_conn = _set_connection_autocommit(_acquire_healthy_pooled_conn(pool))
                return PGCompatConnection(raw_conn, pool=pool)
            except Exception as exc:
                last_error = exc
                logger.warning(
                    "PostgreSQL pooled connection attempt %s/2 failed (%s:%s): %s",
                    attempt + 1,
                    DATABASE_URL_KIND,
                    DATABASE_PORT,
                    exc,
                )
                _reset_pg_pool()

        logger.warning(
            "PostgreSQL pool reset edildi; direct connection fallback deneniyor (%s:%s)",
            DATABASE_URL_KIND,
            DATABASE_PORT,
        )

    for attempt in range(2):
        try:
            raw_conn = _set_connection_autocommit(_create_raw_connection())
            return PGCompatConnection(raw_conn)
        except Exception as exc:
            last_error = exc
            logger.warning(
                "PostgreSQL direct connection attempt %s/2 failed (%s:%s): %s",
                attempt + 1,
                DATABASE_URL_KIND,
                DATABASE_PORT,
                exc,
            )

    if last_error:
        raise last_error
    raise RuntimeError("PostgreSQL bağlantısı kurulamadı")


def parse_dims(dims_str: str) -> tuple[float | None, float | None]:
    """
    Child_Dims string'ini parse eder.
    Örnek: '(49, 63)' → (49.0, 63.0)
    """
    if _is_blank_or_nan(dims_str):
        return None, None
    match = re.match(r"\((\d+(?:\.\d+)?),\s*(\d+(?:\.\d+)?)\)", str(dims_str))
    if match:
        return float(match.group(1)), float(match.group(2))
    return None, None


def _is_blank_or_nan(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float):
        try:
            if math.isnan(value):
                return True
        except Exception:
            pass
    text = str(value).strip()
    if not text:
        return True
    return text.lower() in {"nan", "none", "null"}


def calculate_alan(en: float | None, boy: float | None) -> float | None:
    """Alan = En × Boy / 10000 (m² cinsinden)"""
    if en is not None and boy is not None:
        return round(en * boy / 10000, 6)
    return None


def first_non_empty(*values):
    """İlk dolu (None/boş olmayan) değeri döndürür."""
    for v in values:
        if _is_blank_or_nan(v):
            continue
        return v
    return None


def ensure_products_columns(cursor):
    """Mevcut products tablosuna yeni kolonları migrasyonla ekler."""
    for col_name, col_type in PRODUCT_EXTRA_COLUMNS:
        pg_type = "DOUBLE PRECISION" if col_type.upper() == "REAL" else col_type
        cursor.execute(f"ALTER TABLE products ADD COLUMN IF NOT EXISTS {col_name} {pg_type}")


def backfill_product_raw_cost_status(cursor):
    """
    Eski veritabanlarında yeni eklenen ham maliyet statüsünü ilk kez doldurur.
    Manuel seçimler null olmayacağı için tekrar overwrite edilmez.
    """
    cursor.execute(
        """
        UPDATE products AS p
        SET ham_maliyet_status = CASE
            WHEN EXISTS (
                SELECT 1
                FROM product_materials pm
                WHERE pm.child_sku = p.child_sku
                  AND COALESCE(pm.quantity, 0) > 0
            ) THEN 'calisildi'
            WHEN EXISTS (
                SELECT 1
                FROM product_cost_breakdowns pcb
                WHERE pcb.child_sku = p.child_sku
            ) THEN 'calisildi'
            WHEN COALESCE(TRIM(p.kargo_kodu), '') <> '' THEN 'calisildi'
            WHEN p.kargo_agirlik IS NOT NULL OR p.kargo_desi IS NOT NULL THEN 'calisildi'
            ELSE 'calisilmadi'
        END
        WHERE COALESCE(TRIM(p.ham_maliyet_status), '') = ''
        """
    )


def ensure_audit_logs_columns(cursor):
    """audit_logs tablosuna izlenebilirlik kolonlarını migrasyonla ekler."""
    columns = [
        ("request_id", "TEXT"),
        ("method", "TEXT"),
        ("path", "TEXT"),
        ("ip_address", "TEXT"),
        ("user_agent", "TEXT"),
        ("status", "TEXT"),
    ]
    for col_name, col_type in columns:
        cursor.execute(f"ALTER TABLE audit_logs ADD COLUMN IF NOT EXISTS {col_name} {col_type}")


def ensure_approval_requests_columns(cursor, ref_id_type: str):
    """approval_requests tablosuna workflow kolonlarini migrasyonla ekler."""
    columns = [
        ("request_type", "TEXT"),
        ("target", "TEXT"),
        ("payload", "TEXT"),
        ("status", "TEXT"),
        ("requested_by", ref_id_type),
        ("requested_username", "TEXT"),
        ("reviewed_by", ref_id_type),
        ("reviewed_username", "TEXT"),
        ("review_note", "TEXT"),
        ("execution_result", "TEXT"),
        ("created_at", "TIMESTAMP"),
        ("reviewed_at", "TIMESTAMP"),
        ("executed_at", "TIMESTAMP"),
    ]
    for col_name, col_type in columns:
        cursor.execute(f"ALTER TABLE approval_requests ADD COLUMN IF NOT EXISTS {col_name} {col_type}")


def ensure_indexes(cursor):
    """Sık erişilen sorgular için index'leri oluşturur."""
    index_sql = [
        "CREATE INDEX IF NOT EXISTS idx_products_kategori ON products(kategori)",
        "CREATE INDEX IF NOT EXISTS idx_products_is_active ON products(is_active)",
        "CREATE INDEX IF NOT EXISTS idx_products_parent_id ON products(parent_id)",
        "CREATE INDEX IF NOT EXISTS idx_products_parent_id_active ON products(parent_id, is_active)",
        "CREATE INDEX IF NOT EXISTS idx_products_parent_name ON products(parent_name)",
        "CREATE INDEX IF NOT EXISTS idx_products_parent_kategori ON products(parent_name, kategori)",
        "CREATE INDEX IF NOT EXISTS idx_products_parent_active ON products(parent_name, is_active)",
        "CREATE INDEX IF NOT EXISTS idx_products_active_parent_sku ON products(is_active, parent_name, child_sku)",
        "CREATE INDEX IF NOT EXISTS idx_products_active_kategori_parent ON products(is_active, kategori, parent_name)",
        "CREATE INDEX IF NOT EXISTS idx_products_active_kategori_identifier_sku ON products(is_active, kategori, product_identifier, child_sku)",
        "CREATE INDEX IF NOT EXISTS idx_products_identifier ON products(product_identifier)",
        "CREATE INDEX IF NOT EXISTS idx_products_variation_size ON products(variation_size)",
        "CREATE INDEX IF NOT EXISTS idx_product_materials_child_sku ON product_materials(child_sku)",
        "CREATE INDEX IF NOT EXISTS idx_product_materials_material_id ON product_materials(material_id)",
        "CREATE INDEX IF NOT EXISTS idx_product_costs_child_sku ON product_costs(child_sku)",
        "CREATE INDEX IF NOT EXISTS idx_product_costs_cost_name ON product_costs(cost_name)",
        "CREATE INDEX IF NOT EXISTS idx_product_costs_assigned ON product_costs(assigned)",
        "CREATE INDEX IF NOT EXISTS idx_cost_definitions_category_active ON cost_definitions(category, is_active)",
        "CREATE INDEX IF NOT EXISTS idx_cost_definitions_kargo_code ON cost_definitions(kargo_code)",
        "CREATE INDEX IF NOT EXISTS idx_parent_cost_profiles_parent_id ON parent_cost_profiles(parent_id)",
        "CREATE INDEX IF NOT EXISTS idx_product_cost_breakdowns_parent_id ON product_cost_breakdowns(parent_id)",
        "CREATE INDEX IF NOT EXISTS idx_product_cost_breakdowns_child_sku ON product_cost_breakdowns(child_sku)",
        "CREATE INDEX IF NOT EXISTS idx_parent_cost_groups_name ON parent_cost_groups(name)",
        "CREATE INDEX IF NOT EXISTS idx_parent_cost_group_items_group_id ON parent_cost_group_items(group_id)",
        "CREATE INDEX IF NOT EXISTS idx_parent_cost_group_items_parent_name ON parent_cost_group_items(parent_name)",
        "CREATE INDEX IF NOT EXISTS idx_users_role_active ON users(role, is_active)",
        "CREATE INDEX IF NOT EXISTS idx_users_username_lower ON users(LOWER(username))",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_created_at ON audit_logs(created_at)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_action ON audit_logs(action)",
        "CREATE INDEX IF NOT EXISTS idx_audit_logs_request_id ON audit_logs(request_id)",
        "CREATE INDEX IF NOT EXISTS idx_approval_requests_status_created_at ON approval_requests(status, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_approval_requests_requested_by ON approval_requests(requested_by)",
        "CREATE INDEX IF NOT EXISTS idx_approval_requests_type_status ON approval_requests(request_type, status)",
    ]
    for sql in index_sql:
        cursor.execute(sql)


def init_db():
    """Veritabanını oluşturur ve tabloları hazırlar."""
    conn = get_db()
    cursor = conn.cursor()
    id_col = "BIGSERIAL PRIMARY KEY"
    ref_id_type = "BIGINT"

    # Ürünler tablosu
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS products (
            id {id_col},
            kategori TEXT NOT NULL,
            parent_id REAL,
            parent_name TEXT,
            child_sku TEXT NOT NULL UNIQUE,
            child_name TEXT,
            child_code TEXT,
            child_dims TEXT,
            en REAL,
            boy REAL,
            alan_m2 REAL,
            variation_size TEXT,
            variation_color TEXT,
            product_identifier TEXT,
            match_score REAL,
            match_method TEXT,
            kargo_kodu TEXT,
            kargo_en REAL,
            kargo_boy REAL,
            kargo_yukseklik REAL,
            kargo_agirlik REAL,
            kargo_desi REAL,
            ham_maliyet_status TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    ensure_products_columns(cursor)

    # Hammaddeler tablosu (birim fiyatlar)
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS raw_materials (
            id {id_col},
            name TEXT NOT NULL UNIQUE,
            unit TEXT NOT NULL,
            unit_price REAL NOT NULL DEFAULT 0,
            currency TEXT DEFAULT 'TRY',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Ürün-Hammadde ilişki tablosu (her ürünün hammadde miktarları)
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS product_materials (
            id {id_col},
            child_sku TEXT NOT NULL,
            material_id {ref_id_type} NOT NULL,
            quantity REAL NOT NULL DEFAULT 0,
            FOREIGN KEY (child_sku) REFERENCES products(child_sku),
            FOREIGN KEY (material_id) REFERENCES raw_materials(id),
            UNIQUE(child_sku, material_id)
        )
    """)

    # Maliyet (ambalaj) atamaları tablosu
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS product_costs (
            id {id_col},
            child_sku TEXT NOT NULL,
            cost_name TEXT NOT NULL,
            assigned INTEGER DEFAULT 0,
            FOREIGN KEY (child_sku) REFERENCES products(child_sku),
            UNIQUE(child_sku, cost_name)
        )
    """)

    # Parent seviyesinde tanımlanan maliyet kırılım profilleri
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS parent_cost_profiles (
            id {id_col},
            parent_id REAL NOT NULL UNIQUE,
            parent_name TEXT,
            parent_sku TEXT,
            breakdown_payload TEXT NOT NULL,
            updated_by {ref_id_type},
            updated_by_username TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Child seviyesine aynalanan maliyet kırılımı
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS product_cost_breakdowns (
            id {id_col},
            child_sku TEXT NOT NULL UNIQUE,
            parent_id REAL,
            breakdown_payload TEXT NOT NULL,
            updated_by {ref_id_type},
            updated_by_username TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (child_sku) REFERENCES products(child_sku)
        )
    """)

    # Parent maliyet grupları (birden fazla parent'ı aynı maliyet setinde toplamak için)
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS parent_cost_groups (
            id {id_col},
            name TEXT NOT NULL UNIQUE,
            description TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_by {ref_id_type},
            created_by_username TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Parent maliyet grup üyeleri
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS parent_cost_group_items (
            id {id_col},
            group_id {ref_id_type} NOT NULL,
            parent_name TEXT NOT NULL,
            kategori TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (group_id) REFERENCES parent_cost_groups(id),
            UNIQUE(group_id, parent_name)
        )
    """)

    # Maliyet tanımları (kargo/kaplama) tablosu
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS cost_definitions (
            id {id_col},
            name TEXT NOT NULL UNIQUE,
            category TEXT NOT NULL DEFAULT 'kaplama',
            kargo_code TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            source TEXT DEFAULT 'manual',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Kullanıcılar (rol bazlı yetki)
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS users (
            id {id_col},
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Basit audit kaydı (opsiyonel izleme)
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id {id_col},
            user_id {ref_id_type},
            username TEXT,
            action TEXT NOT NULL,
            target TEXT,
            details TEXT,
            request_id TEXT,
            method TEXT,
            path TEXT,
            ip_address TEXT,
            user_agent TEXT,
            status TEXT DEFAULT 'ok',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    ensure_audit_logs_columns(cursor)

    # Onay workflow kayıtları
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS approval_requests (
            id {id_col},
            request_type TEXT NOT NULL,
            target TEXT,
            payload TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            requested_by {ref_id_type},
            requested_username TEXT,
            reviewed_by {ref_id_type},
            reviewed_username TEXT,
            review_note TEXT,
            execution_result TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            reviewed_at TIMESTAMP,
            executed_at TIMESTAMP
        )
    """)
    ensure_approval_requests_columns(cursor, ref_id_type=ref_id_type)

    # Sorgu performansı için index paketini uygula
    ensure_indexes(cursor)
    backfill_product_raw_cost_status(cursor)

    conn.commit()
    conn.close()


def _resolve_kategori_csv_path(kategori_name: str, kategori_dir: Path) -> Path | None:
    env_key = f"KATEGORI_{kategori_name.upper()}_CSV_PATH"
    env_value = os.getenv(env_key, "").strip()
    candidates: list[Path] = []

    if env_value:
        if is_http_url(env_value):
            cached = cache_remote_file(
                env_value,
                cache_name=f"{kategori_name}_mapped_products.csv",
                ttl_seconds=int(os.getenv("REMOTE_FILE_CACHE_TTL", "900")),
            )
            if cached and cached.exists():
                return cached
        else:
            candidates.append(Path(env_value).expanduser())

    candidates.extend([
        PROJECT_ROOT / f"mapped_products_{kategori_name}.csv",
        PROJECT_ROOT / f"{kategori_name}_kategori_list.csv",
        PROJECT_ROOT / f"{kategori_name}_mapping_result.csv",
        kategori_dir / f"{kategori_name}_kategori_list.csv",
        kategori_dir / "mapped_products.csv",
        kategori_dir / f"mapped_products_{kategori_name}.csv",
        kategori_dir / f"{kategori_name}_mapping_result.csv",
    ])

    if kategori_name == "cam":
        candidates.append(kategori_dir / "glass_mapping_result.csv")
    if kategori_name == "harita":
        candidates.append(kategori_dir / "harita_mapping_result.csv")

    deduped: list[Path] = []
    seen = set()
    for path in candidates:
        key = str(path)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(path)

    return next((p for p in deduped if p.exists()), None)


def load_mapped_products(
    categories: list[str] | None = None,
    replace_existing: bool = False,
    preserved_raw_cost_statuses: dict[str, str] | None = None,
):
    """
    mapped product listelerini DB'ye yükler.
    - categories: sadece seçili kategorileri güncelle
    - replace_existing: true ise seçili kategorilerdeki eski ürünleri temizleyip yeniden yükle
    """
    selected_categories = normalize_product_categories(categories)
    conn = get_db()
    cursor = conn.cursor()
    status_by_sku: dict[str, str] = {}

    if preserved_raw_cost_statuses:
        for child_sku, status in preserved_raw_cost_statuses.items():
            sku = str(child_sku or "").strip()
            normalized_status = str(status or "").strip().lower()
            if not sku or normalized_status not in {"calisildi", "calisilmadi"}:
                continue
            status_by_sku[sku] = normalized_status

    if replace_existing:
        placeholders = ", ".join(["?"] * len(selected_categories))
        existing_status_rows = cursor.execute(
            f"""
            SELECT child_sku, ham_maliyet_status
            FROM products
            WHERE kategori IN ({placeholders})
            """,
            selected_categories,
        ).fetchall()
        for row in existing_status_rows:
            sku = str(row["child_sku"] or "").strip()
            normalized_status = str(row["ham_maliyet_status"] or "").strip().lower()
            if not sku or normalized_status not in {"calisildi", "calisilmadi"}:
                continue
            status_by_sku.setdefault(sku, normalized_status)

        for kategori_name in selected_categories:
            cursor.execute(
                "DELETE FROM product_materials WHERE child_sku IN (SELECT child_sku FROM products WHERE kategori = ?)",
                (kategori_name,),
            )
            cursor.execute(
                "DELETE FROM product_cost_breakdowns WHERE child_sku IN (SELECT child_sku FROM products WHERE kategori = ?)",
                (kategori_name,),
            )
            cursor.execute(
                "DELETE FROM product_costs WHERE child_sku IN (SELECT child_sku FROM products WHERE kategori = ?)",
                (kategori_name,),
            )
            cursor.execute("DELETE FROM products WHERE kategori = ?", (kategori_name,))

    total_loaded = 0

    for kategori_name in selected_categories:
        kategori_dir = KATEGORI_DIRS[kategori_name]
        csv_path = _resolve_kategori_csv_path(kategori_name, kategori_dir)
        if csv_path is None:
            print(f"⚠ {kategori_name}: CSV bulunamadı, atlanıyor.")
            continue

        with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))

        print(f"📦 {kategori_name}: {len(rows)} ürün yükleniyor...")

        for row in rows:
            child_sku = first_non_empty(row.get("Child_SKU"), row.get("Child_ID"))
            if not child_sku:
                continue

            child_name = first_non_empty(row.get("Child_Name"))
            child_code = first_non_empty(row.get("Child_Code"))
            child_dims = first_non_empty(row.get("Child_Dims"))
            variation_size = first_non_empty(row.get("variationSize"), row.get("Variation_Size"))
            variation_color = first_non_empty(row.get("variationColor"), row.get("Variation_Color"))
            product_identifier = first_non_empty(row.get("productIdentifier"), row.get("Product_Identifier"))
            match_score = first_non_empty(row.get("Match_Score"), row.get("Match_Confidence_Score"))
            match_method = first_non_empty(row.get("Match_Method"), row.get("Manual_Review"))

            en, boy = parse_dims(child_dims) if child_dims else (None, None)
            alan = calculate_alan(en, boy)
            is_active = 0 if str(child_code or "").strip().upper().startswith("CUS") else 1
            ham_maliyet_status = status_by_sku.get(child_sku, "calisilmadi")
            values = (
                kategori_name,
                row.get("Parent_ID"),
                row.get("Parent_Name"),
                child_sku,
                child_name,
                child_code,
                child_dims,
                en,
                boy,
                alan,
                variation_size,
                variation_color,
                product_identifier,
                match_score,
                match_method,
                ham_maliyet_status,
                is_active,
            )

            try:
                cursor.execute("""
                    INSERT INTO products
                    (kategori, parent_id, parent_name, child_sku, child_name,
                     child_code, child_dims, en, boy, alan_m2,
                     variation_size, variation_color, product_identifier,
                     match_score, match_method, ham_maliyet_status, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT (child_sku) DO UPDATE SET
                        kategori = EXCLUDED.kategori,
                        parent_id = EXCLUDED.parent_id,
                        parent_name = EXCLUDED.parent_name,
                        child_name = EXCLUDED.child_name,
                        child_code = EXCLUDED.child_code,
                        child_dims = EXCLUDED.child_dims,
                        en = EXCLUDED.en,
                        boy = EXCLUDED.boy,
                        alan_m2 = EXCLUDED.alan_m2,
                        variation_size = EXCLUDED.variation_size,
                        variation_color = EXCLUDED.variation_color,
                        product_identifier = EXCLUDED.product_identifier,
                        match_score = EXCLUDED.match_score,
                        match_method = EXCLUDED.match_method,
                        is_active = EXCLUDED.is_active
                """, values)
                total_loaded += 1
            except Exception as e:
                print(f"  ⚠ Hata ({child_sku}): {e}")

    conn.commit()
    conn.close()
    print(f"✅ Toplam {total_loaded} ürün yüklendi.")
    return total_loaded


def deactivate_cus_products() -> int:
    """
    Child_Code 'CUS' ile başlayan ürünleri pasife alır.
    """
    conn = get_db()
    cur = conn.execute(
        """
        UPDATE products
        SET is_active = 0
        WHERE UPPER(TRIM(COALESCE(child_code, ''))) LIKE ?
          AND COALESCE(is_active, 1) <> 0
        """,
        ("CUS%",),
    )
    conn.commit()
    affected = int(cur.rowcount or 0) if (cur.rowcount or 0) > 0 else 0
    conn.close()
    return affected


def load_default_materials():
    """maliyet_sablonu.xlsx'ten varsayılan hammadde listesini yükler."""
    import openpyxl

    template_path = resolve_template_path()
    if not template_path.exists():
        print(f"⚠ {template_path} bulunamadı")
        return

    wb = openpyxl.load_workbook(template_path, read_only=True)
    ws = wb["Maliyet Şablonu"]

    conn = get_db()
    cursor = conn.cursor()

    # Hammadde kolonlarını bul (DX-EX arası, "Hammadde:" ile başlayan)
    material_count = 0
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).startswith("Hammadde:"):
            # "Hammadde: UV (m2)" → name="UV", unit="m2"
            raw = str(header).replace("Hammadde:", "").strip()
            # Birim parantez içinde
            unit_match = re.search(r"\(([^)]+)\)", raw)
            unit = unit_match.group(1) if unit_match else "pcs"
            name = re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()

            try:
                cursor.execute("""
                    INSERT OR IGNORE INTO raw_materials (name, unit, unit_price)
                    VALUES (?, ?, 0)
                """, (name, unit))
                if int(getattr(cursor, "rowcount", 0) or 0) > 0:
                    material_count += 1
            except Exception as e:
                print(f"  ⚠ Hammadde yüklenirken hata ({name}): {e}")

    conn.commit()
    conn.close()
    wb.close()
    print(f"✅ {material_count} hammadde tanımı yüklendi.")
    return material_count


def extract_kargo_code_from_name(name: str | None) -> str | None:
    """Maliyet adından M-8 gibi kargo kodunu normalize ederek çıkarır."""
    if not name:
        return None
    m = KARGO_CODE_PATTERN.search(str(name).upper())
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}"


def load_template_cost_names():
    """maliyet_sablonu.xlsx'ten maliyet (ambalaj) isimlerini okur."""
    import openpyxl

    template_path = resolve_template_path()
    if not template_path.exists():
        return []

    wb = openpyxl.load_workbook(template_path, read_only=True)
    ws = wb["Maliyet Şablonu"]

    cost_names = []
    seen: set[str] = set()
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).startswith("Maliyet:"):
            name = str(header).replace("Maliyet:", "").strip()
            name = canonicalize_kaplama_cost_name(name)
            if not name or name in seen:
                continue
            seen.add(name)
            cost_names.append(name)

    wb.close()
    return cost_names


def sync_cost_definitions_from_template():
    """
    Şablondaki maliyet isimlerini cost_definitions tablosuna ekler.
    Var olan kayıtları ezmez, sadece eksikleri INSERT eder.
    """
    cost_names = load_template_cost_names()
    if not cost_names:
        return 0

    conn = get_db()
    cursor = conn.cursor()
    inserted = 0

    for name in cost_names:
        code = extract_kargo_code_from_name(name)
        category = "kargo" if code else "kaplama"
        cursor.execute("""
            INSERT OR IGNORE INTO cost_definitions
            (name, category, kargo_code, is_active, source)
            VALUES (?, ?, ?, 1, 'template')
        """, (name, category, code))
        if cursor.rowcount:
            inserted += 1

    conn.commit()
    conn.close()
    return inserted


def normalize_legacy_gold_silver_names() -> int:
    """
    Legacy kaplama adlarını '(gold,copper)' formatına çevirir.
    - cost_definitions kayıtlarını birleştirir
    - product_costs referanslarını çakışma oluşturmadan taşır
    """
    conn = get_db()
    cursor = conn.cursor()

    rows = cursor.execute("""
        SELECT id, name
        FROM cost_definitions
        WHERE LOWER(name) LIKE '%(gold,silver)%'
    """).fetchall()

    changed = 0

    for row in rows:
        old_name = str(row["name"] or "").strip()
        new_name = canonicalize_kaplama_cost_name(old_name)
        if not old_name or not new_name or old_name == new_name:
            continue

        # 1) product_costs referanslarını yeni ada taşı
        cursor.execute("""
            UPDATE product_costs
            SET cost_name = ?
            WHERE cost_name = ?
              AND NOT EXISTS (
                    SELECT 1
                    FROM product_costs pc2
                    WHERE pc2.child_sku = product_costs.child_sku
                      AND pc2.cost_name = ?
              )
        """, (new_name, old_name, new_name))

        # Hedef ad zaten varsa eski duplicate kayıtları temizle
        cursor.execute("""
            DELETE FROM product_costs
            WHERE cost_name = ?
              AND EXISTS (
                    SELECT 1
                    FROM product_costs pc2
                    WHERE pc2.child_sku = product_costs.child_sku
                      AND pc2.cost_name = ?
              )
        """, (old_name, new_name))

        # 2) cost_definitions tarafını birleştir
        existing = cursor.execute(
            "SELECT id, is_active FROM cost_definitions WHERE name = ?",
            (new_name,),
        ).fetchone()

        if existing:
            if int(existing["is_active"] or 0) == 0:
                cursor.execute(
                    "UPDATE cost_definitions SET is_active = 1, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (existing["id"],),
                )
            cursor.execute("DELETE FROM cost_definitions WHERE id = ?", (row["id"],))
        else:
            cursor.execute(
                "UPDATE cost_definitions SET name = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (new_name, row["id"]),
            )

        changed += 1

    # cost_definitions kaydı olmasa bile product_costs içinde legacy ad kalmış olabilir
    legacy_product_costs = cursor.execute("""
        SELECT DISTINCT cost_name
        FROM product_costs
        WHERE LOWER(cost_name) LIKE '%(gold,silver)%'
    """).fetchall()

    for row in legacy_product_costs:
        old_name = str(row["cost_name"] or "").strip()
        new_name = canonicalize_kaplama_cost_name(old_name)
        if not old_name or not new_name or old_name == new_name:
            continue

        cursor.execute("""
            UPDATE product_costs
            SET cost_name = ?
            WHERE cost_name = ?
              AND NOT EXISTS (
                    SELECT 1
                    FROM product_costs pc2
                    WHERE pc2.child_sku = product_costs.child_sku
                      AND pc2.cost_name = ?
              )
        """, (new_name, old_name, new_name))

        cursor.execute("""
            DELETE FROM product_costs
            WHERE cost_name = ?
              AND EXISTS (
                    SELECT 1
                    FROM product_costs pc2
                    WHERE pc2.child_sku = product_costs.child_sku
                      AND pc2.cost_name = ?
              )
        """, (old_name, new_name))

        code = extract_kargo_code_from_name(new_name)
        category = "kargo" if code else "kaplama"
        cursor.execute("""
            INSERT OR IGNORE INTO cost_definitions
            (name, category, kargo_code, is_active, source)
            VALUES (?, ?, ?, 1, 'legacy_migration')
        """, (new_name, category, code))

        changed += 1

    conn.commit()
    conn.close()
    return changed


def deactivate_shadowed_kaplama_base_names() -> int:
    """
    Hem tier'lı (silver/gold,copper) hem de suffix'siz hali bulunan kaplama adlarında
    suffix'siz olan eski kayıtları pasife alır.
    """
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT id, name, is_active
        FROM cost_definitions
        WHERE category = 'kaplama'
    """).fetchall()

    tiered_base_keys: set[str] = set()
    flat_rows: list = []

    for row in rows:
        name = str(row["name"] or "").strip()
        base, suffix = split_kaplama_tier_suffix(name)
        key = base.casefold()
        if not key:
            continue
        if suffix:
            tiered_base_keys.add(key)
        else:
            flat_rows.append(row)

    deactivate_ids = [
        int(row["id"])
        for row in flat_rows
        if str(row["name"] or "").strip().casefold() in tiered_base_keys and int(row["is_active"] or 0) == 1
    ]

    for cost_id in deactivate_ids:
        cursor.execute(
            "UPDATE cost_definitions SET is_active = 0, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (cost_id,),
        )

    conn.commit()
    conn.close()
    return len(deactivate_ids)


def list_cost_definitions(active_only: bool = True, category: str | None = None):
    """Maliyet tanımlarını döndürür."""
    conn = get_db()

    where = []
    params: list = []
    if active_only:
        where.append("is_active = 1")
    if category:
        where.append("category = ?")
        params.append(category)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    rows = conn.execute(f"""
        SELECT id, name, category, kargo_code, is_active, source, created_at, updated_at
        FROM cost_definitions
        {where_sql}
        ORDER BY category, name
    """, params).fetchall()

    if not rows:
        conn.close()
        # Template'den sync edip tekrar dene
        sync_cost_definitions_from_template()
        normalize_legacy_gold_silver_names()
        deactivate_shadowed_kaplama_base_names()
        conn = get_db()
        rows = conn.execute(f"""
            SELECT id, name, category, kargo_code, is_active, source, created_at, updated_at
            FROM cost_definitions
            {where_sql}
            ORDER BY category, name
        """, params).fetchall()

    result = [dict(r) for r in rows]
    conn.close()
    return result


def load_cost_names():
    """Aktif maliyet isimlerini cost_definitions tablosundan döndürür."""
    defs = list_cost_definitions(active_only=True)
    return [d["name"] for d in defs]


if __name__ == "__main__":
    print("🔧 Veritabanı başlatılıyor...")
    init_db()
    print("📦 Mapped products yükleniyor...")
    load_mapped_products()
    print("🧪 Hammaddeler yükleniyor...")
    load_default_materials()
    print("✅ Veritabanı hazır!")
